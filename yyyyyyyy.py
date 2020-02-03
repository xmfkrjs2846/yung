from bs4 import BeautifulSoup as bs
import requests
import discord
import time
import random
import openpyxl
import os
import asyncio

client = discord.Client()
talk = [0]

@client.event
async def on_ready():
    print(client.user.id)
    print("ready")
    await client.change_presence(game=discord.Game(name="버그시 호출;;| help", type=1))
@client.event
async def on_member_join(member):
    fmt = '[{1.name}] 에 입사한걸 환영해~, {0.mention}아'
    channel = member.server.get_channel("527835129150832641")
    await client.send_message(channel, fmt.format(member, member.server))
    await client.send_message(channel,
                              embed=discord.Embed(colour=discord.Colour.gold(), title='신입?! 이런 채널에?? 헐....;;'))
    time.sleep(1)
    await client.send_message(channel, "나는 기능하고 회화적인 부분을 도와줄꺼야")
    time.sleep(1)
    await client.send_message(channel, "아오바는 음악적인 부분을 도와줄꺼야")

@client.event
async def on_member_remove(member):
    channel = member.server.get_channel("527835129150832641")
    fmt = '{0.mention}가 퇴사했어! 정말 현명한 선택인 것 같아.'
    await client.send_message(channel, fmt.format(member, member.server))

@client.event
async def on_message(message):
 if message.author != client.user:
   if message.content.startswith("퇴근"):
         await client.send_message(message.channel, "수고하셨어요~! 저 먼저 퇴근할게요!\n(앞으로 융이 아무 말도 하지 않습니다)")
         del talk[0]
         talk.append(1)
         await client.change_presence(status=discord.Status.idle, game=discord.Game(name="집에서 노는 것을", type=1))
   if message.content.startswith("출근"):
         await client.send_message(message.channel, "오늘도 야근인가... 후우...;;")
         del talk[0]
         talk.append(0)
         await client.change_presence(game=discord.Game(name="버그시 호출;;| help", type=1))
   if talk[0] == 0:
    if message.content.startswith("융"):
        await client.send_message(message.channel,"융융!")
    if message.content.startswith("혹시"):
        await client.send_file(message.channel,'혹시.gif')
        await client.send_message(message.channel,"흐.흐흐..!")
    if message.content.startswith("ㅋ"):
        await client.send_message(message.channel,"ㅋㅋㅋㅋㅋㅋ")
    if message.content.startswith("ㅗ"):
        await client.send_message(message.channel,"빡큐")
        time.sleep(2)
        await client.send_message(message.channel, "뽀큐! 뽁큐!")
        await client.send_message(message.channel, "빡큐 ㅗ^오^ㅗ")

    if message.content.startswith("야"):
        await client.send_message(message.channel,"응?! 나 불렀어??!")
        time.sleep(1)
        await client.send_message(message.channel,"미안.. 관심좀 끌어봤어;;")
    if message.content.startswith("에픽세븐") or message.content.startswith("-에"):
        await client.send_message(message.channel,"1320만?! 내 월급의 4,5배 잖아?!!")
        time.sleep(1)
        await client.send_message(message.channel,"뭐지 이 낯선 천장은... 후우...")
    if message.content.startswith("https"):
        await client.send_message(message.channel,"머야?머야? 나도 볼래!!")
    if message.content.startswith("Us:track"):
        await client.send_message(message.channel,"코이카케루신아이카노조!")
    if message.content.startswith("집가고싶다"):
        await client.send_message(message.channel, "ㅋㅋㅋㅋ 나는 평생 노는데 부럽지? 부러럽지?? 부럽지???!")
    if message.content.startswith("심심"):
        await client.send_message(message.channel,"휴가 나왔어? 부럽네~")
    if message.content.startswith("ㅇㅈ") or message.content.startswith("인정"):
        injun = ["저기 이건 내가 생각하기에 좀 아니라도 생각하는데...",
                 "너 의외로 맞는 말 한다?!"]
        hey1 = random.choice(injun)
        await client.send_message(message.channel, "{}".format(hey1))
    if message.content.startswith("ㄷㅊ") or message.content.startswith("닥쳐"):
        await client.send_message(message.channel, "선배 죄송합니다...ㅜㅜ 다시 일 시작할게요ㅠㅠ 해고만은!!")
    if message.content.startswith("융털"):
        await client.send_file(message.channel, '11111.gif')
        await client.send_message(message.channel, "벼...변.태..")
    if message.content.startswith("바보"):
        await client.send_file(message.channel, '바보.jpg')
        await client.send_message(message.channel,"너한테 바보라는데? 풉!")
    if message.content.startswith("이이잉"):
        await client.send_message(message.channel, "앗살라말라이쿰")
        await client.send_file(message.channel, '안경.png')
    if message.content.startswith("섹스"):
        await client.send_file(message.channel, '나 살쪘나...gif')
        await client.send_message(message.channel, "에!..엣.. 헤..헨따이!!")
    if message.content.startswith("시발") or message.content.startswith("ㅅㅂ"):
        await client.send_message(message.channel, "나쁜말? 나쁜말?! ㅡ.ㅡ")
    if message.content.startswith("ㄹㅇ"):
        await client.send_message(message.channel, "너어어..는 저어어어엉.말")
    if message.content.startswith("휴"):
        await client.send_message(message.channel, "뭔가 안좋은일이 라도 있었어??")
    if message.content.startswith("아님") or message.content.startswith("ㄴㄴ") or message.content.startswith("아니"):
        await client.send_file(message.channel, '아 아니야.gif')
        await client.send_message(message.channel, "머가 아닌데?? 솔직히 말해바~ 내가 다 들어줄게~ㅋㅋ")
    if message.content.startswith("메이플"):
        if message.content[0:] =="메이플":
            await client.send_file(message.channel,'일.gif')
            await client.send_message(message.channel, "일도 좋지만 좀 쉬어가면서 하는게 어때??!")
    if message.content.startswith("뜨끔") or message.content.startswith("들킴"):
        await client.send_file(message.channel, '두리번 두리번.gif')
        await client.send_message(message.channel, "....?")
    if message.content.startswith("병신") or message.content.startswith("ㅂㅅ") or message.content.startswith("ㅄ"):
        await client.send_message(message.channel,"....???")
        time.sleep(1)
        await client.send_message(message.channel,"아! 뭔지 알았어~! ㅂㅅ인지 검증해볼게~!")
        time.sleep(2)
        answer1 = ["삐비빅 넌 몸이 아픈걸로 판정났어! 병원이라도 가봐~!",
                  "칫! 널 바보 취급 할 수 있는 기회였는데 정상이라니..."]
        hey = random.choice(answer1)
        await client.send_message(message.channel,"{}".format(hey))
    if message.content.startswith("역겹") or message.content.startswith("ㄴㄷㅆ") or message.content.startswith("네다씹") or message.content.startswith("무서"):
        await client.send_file(message.channel, '무서.jpg')
        await client.send_message(message.channel, "나랑 3M만 떨어줘줄래? 조끔 그렇다..")
    if message.content.startswith("힘내") or message.content.startswith("화이팅"):
        await client.send_message(message.channel, "아자아자!!")
        await client.send_file(message.channel, '아자.jpg')
    #호출
    if message.content.startswith('준혁'):
        myid = '<@379966497293991937>'
        me = await client.get_user_info('379966497293991937')
        msg = "{0.author.mention}가 너를 부르고 있어~ 빨리 가보는게 어때?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_file(message.channel, '뜨끔...gif')
        await client.send_message(message.channel,"{}누군가 널 부르는거 같은데??".format(myid))
    if message.content.startswith('태훈'):
        myid = '<@344788669502193666>'
        me = await client.get_user_info('344788669502193666')
        msg = "{0.author.mention}가 너를 부르고 있어~ 빨리 가보는게 어때?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel,"{}누군가 널 부르는거 같은데??".format(myid))
    if message.content.startswith('대영'):
        myid='<@368617760412139520>'
        me = await client.get_user_info('368617760412139520')
        msg = "{0.author.mention}상가 키미또 엣찌이나 코토 시타이떼 이떼이루요!!".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}아다 누가 너 부른다잉 씨게 대답 하그라!".format(myid))
    if message.content.startswith('영웅'):
        myid='<@286329350633029634>'
        me = await client.get_user_info('286329350633029634')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}169.7cm..?그사람 여자야? 170cm가 안된다고?!".format(myid))
    if message.content.startswith('후석'):
        myid='<@247651944926019584>'
        me = await client.get_user_info('247651944926019584')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}저..저기..어떡해 이사람 꼭 불러야해??ㅠㅠ".format(myid))
    if message.content.startswith('종민'):
        myid = '<@206663386937688065>'
        me = await client.get_user_info('206663386937688065')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}리듬게임하고 있는 것 같은데 잠깐 나랑 이야기좀 해줄 수 있니?".format(myid))
    if message.content.startswith('민기'):
        myid = '<@253749865035726848>'
        me = await client.get_user_info('253749865035726848')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}저기! 당신의 이름이 뜻하는건 뭔지 알려주세요!!".format(myid))
    if message.content.startswith('권영'):
        myid = '<@410450620823371778>'
        me = await client.get_user_info('410450620823371778')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}さんの共に마인크라프토しませんか".format(myid))
    if message.content.startswith('현개') or message.content.startswith("성현"):
        myid = '<@245857392623878146>'
        me = await client.get_user_info('245857392623878146')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}연꽃? 렌게?! 흐음~ 뭔가 떠오르는 것 같기도 하고 아무튼 당신을 부르고있어요.".format(myid))
    if message.content.startswith('종원'):
        myid = '<@271957069991641088>'
        me = await client.get_user_info('271957069991641088')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}아이돌마스터?? 그게 머야? 먹는거야? 그리고 혹시 러브라이브 좋아해??!ㅋ".format(myid))
    if message.content.startswith('희태'):
        myid = '<@241906227611697153>'
        me = await client.get_user_info('241906227611697153')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}벌레다!! 꺅!.. 서..선배.. 저 벌레좀 어떻게 해주세요!ㅠㅠ ".format(myid))
    if message.content.startswith('성규') or message.content.startswith('슬기') or message.content.startswith('성구'):
        myid = '<@356311092332593152>'
        me = await client.get_user_info('356311092332593152')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{} 438만원 쓴 음머어~ 찾아요!!".format(myid))
    if message.content.startswith("안녕"):
        if message.content[0:] == '안녕':
             msg = "반가워~ {0.author.mention}아!!".format(message)
             await client.send_file(message.channel,'새로운 안녕.gif')
             await client.send_message(message.channel,"{}".format(msg))
    if message.content.startswith('강성'):
        myid = '<@386501875379470338>'
        me = await client.get_user_info('386501875379470338')
        msg = "{0.author.mention}이가 부르고있어 어서 대답해주는게 좋을 것 같은데?".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}저기 메이플 대체 언제 그만 두실 꺼에요?!".format(myid))
    if message.content.startswith('흑인'):
        myid = '<@386501875379470338>'
        me = await client.get_user_info('386501875379470338')
        msg = "{0.author.mention}흑흑흗흑흑흑흑흑흫흑흑ㅎ긓긓ㄱ흐긓긓ㄱ".format(message)
        await client.send_message(me, "{}".format(msg))
        await client.send_message(message.channel, "{}저기 메이플 대체 언제 그만 두실 꺼에요?!".format(myid))
    if message.content.startswith("안녕하살법"):
        await client.send_file(message.channel, '받아치기.gif')
        await client.send_message(message.channel,embed=discord.Embed(colour=discord.Colour.gold(), title='안녕하살법 받아치기!'))
    if message.content.startswith("!융"):
        msgg=message.content[3:]
        answer = [
        '그게 좋은거같아!',
        '나 그거 완전 좋아',
        '그게 먼데 난 처음듣는건데??',
        '흐헤헤 이거 치워줘 ㅜㅜ',
        '넌 미움받을 짓만 하는구나..',
        '아무생각 없다',
        '(딴청을 피우는 것 같다']
        total=random.choice(answer)
        people='{0.author.mention}'.format(message)
        await client.send_message(message.channel,'{}의 질문:{}\n융의 대답:{}'.format(people,msgg,total))
    if message.content.startswith('날씨'):
        search=message.content[3:]
        html = requests.get('https://search.naver.com/search.naver?query={} 날씨'.format(search))
        soup = bs(html.text, 'html.parser')

        data1 = soup.find('ul', {'class': 'info_list'})
        data3 = data1.find('p', {'class': 'cast_txt'}).text
        data2 = soup.find('p', {'class': 'info_temperature'})
        data4 = data2.find('span',{'class':'todaytemp'}).text

        data5 = soup.find('div', {'class': 'tomorrow_area'})
        data6 = data5.find('p', {'class': 'info_temperature'})
        data8 = data6.find('span', {'class': 'todaytemp'}).text
        data7 = data5.find('ul', {'class': 'info_list'}).text

        tomorrowAreaBase = soup.find('div', {'class': 'tomorrow_area'})
        tomorrowAllFind = tomorrowAreaBase.find_all('div', {'class': 'main_info morning_box'})
        tomorrowAfter1 = tomorrowAllFind[1]
        tomorrowAfter2 = tomorrowAfter1.find('p', {'class': 'info_temperature'})
        tomorrowAfter3 = tomorrowAfter2.find('span', {'class': 'todaytemp'})
        tomorrowAfterTemp = tomorrowAfter3.text.strip()   #온도
        tomorrowAfterValue1 = tomorrowAfter1.find('div', {'class': 'info_data'})
        tomorrowAfterValue = tomorrowAfterValue1.text.strip() #날씨
        await client.send_message(message.channel, '현재 {}의 온도는 {}˚이며\n날씨는 {}'.format(search,data4,data3))
        await client.send_message(message.channel, '--------------------------\n내일 오전의 온도는 {}˚이며\n날씨는 {}'.format(data8,data7))
        await client.send_message(message.channel,'--------------------------\n내일 오후의 온도는 {}˚이며\n날씨는 {}'.format(tomorrowAfterTemp,tomorrowAfterValue))
        time.sleep(1)
        await client.send_message(message.channel,'-----------------------\n이상 이글점프소속 그래피커 이이지마 윤이였습니다!!')

    if message.content.startswith("골라"):
            choice = message.content.split(" ")
            chonumber = random.randint(1, len(choice) - 1)
            choiceresult = choice[chonumber]
            await client.send_message(message.channel, '나는 "{}" 좋다고 생각해!!'.format(choiceresult))


    if message.content.startswith("주사위"):
            dice = random.randrange(1, 6)
            await client.send_message(message.channel, """너한테는 불행이 어울려보여~! 내가 3초 세줄게! 3초!""")
            time.sleep(1)
            await client.send_message(message.channel, "2초!")
            time.sleep(1)
            await client.send_message(message.channel, "1초!")
            time.sleep(1)
            await client.send_message(message.channel, "짜짠! 숫자 '{}이 나왔습니다!!".format(dice))

    if message.content.startswith('청소'):
            den = int(message.content[3:])
            if den > 0:
                async for m in client.logs_from(message.channel, limit=den + 1):
                    await client.delete_message(m)
                await client.send_message(message.channel, " 메세지 {}개를 청소했어! 잘했지?!".format(str(den)))
            else:
                await client.send_message(message.channel, "1개 이상의 메시지만 삭제할 수 있습니다!")

    if message.content.startswith('타이머'):
            Text = message.content[4:]
            sec = int(Text)
            for i in range(sec, 0, -1):
                await client.send_message(message.channel, embed=discord.Embed(description='타이머 작동중 : ' + str(i) + '초'))
                time.sleep(1)
            else:
                await client.send_message(message.channel,
                                          embed=discord.Embed(description='땡땡땡', colour=discord.Colour.gold()))
    if message.content.startswith('저장'):
       file = openpyxl.load_workbook('저장.xlsx')
       sheet = file.active
       learn = message.content.split(" ")
       for i in range(1, 201):
               if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                    sheet["A" + str(i)].value = learn[1]
                    sheet["B" + str(i)].value = learn[2]
                    await client.send_message(message.channel, "말해 {}이라고 하면 내가 {}이라고 앞으로 대답해줄게".format(learn[1],learn[2]))
                    await client.send_message(message.channel, "★ 현재 사용중인 데이터 저장용량 : 200/" + str(i) + " ★")
                    break
       file.save("저장.xlsx")
    if message.content.startswith("말해"):
         file = openpyxl.load_workbook("저장.xlsx")
         sheet = file.active
         memory = message.content.split(" ")
         for i in range(1, 201):
                  if sheet["A" + str(i)].value == memory[1]:
                        await client.send_message(message.channel, sheet["B" + str(i)].value)
                        break
    if message.content.startswith("기억데이터"):
         file = openpyxl.load_workbook("저장.xlsx")
         sheet = file.active
         for i in range(1, 201):
                   if sheet["A" + str(i)].value == "-" and i == 1:
                          await client.send_message(message.channel, "데이터 없음")
                   if sheet["A" + str(i)].value == "-":
                           break
                   await client.send_message(message.channel,
                                   "A : " + sheet["A" + str(i)].value + " B : " + sheet["B" + str(i)].value)
    if message.content.startswith("조용"):
        role = "에엣취"
        name2 = message.content.split(" ")
        name3 = name2[1]
        name4 = name3[3:21]
        print(name4)

        member = discord.utils.get(client.get_all_members(), id=name4)
        for i in message.server.roles:
            if i.name == "에엣취":
                role = i
                break
        await client.add_roles(member, role)
        for i in message.server.roles:
            if i.name == "메이플안하는사람":
                role = i
                break
        await client.remove_roles(member, role)
        await client.send_message(message.channel, "팀장의 명령때문에 {}씨를 벙어리로 만들었어요~ ㅠㅠ 죄송합니다!!".format(name3))
    if message.content.startswith("취소"):
        role = "에엣취"
        name2 = message.content.split(" ")
        name3 = name2[1]
        name4 = name3[3:21]
        print(name4)

        member = discord.utils.get(client.get_all_members(), id=name4)
        for i in message.server.roles:
            if i.name == "에엣취":
                role = i
                break
        await client.remove_roles(member, role)
        for i in message.server.roles:
            if i.name == "메이플안하는사람":
                role = i
                break
        await client.add_roles(member, role)
        await client.send_message(message.channel, "{}야 너 징계 풀어준데 잘됐지?! 그치그치??".format(name3))

    if message.content.startswith('검색'):
        life = message.content[3:]
        learn = life.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        encText= learn[0]
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
             encText=('{}+{}'.format(encText,learn[i]))
        html1 = ('https://search.naver.com/search.naver?query={}'.format(encText))
        html4 =('https://www.youtube.com/results?search_query={}'.format(encText))
        html2 =('https://www.google.com/search?biw=1920&bih=969&ei=2pClXdu9DdXemAWLkKyoDw&q={}&oq={}&gs_l=psy-ab.3..0l10.60853.65724..65879...12.0..3.219.2223.8j11j1......0....1..gws-wiz.....0..0i67j0i131j0i10j0i10i30j0i5i10i30.EV6bS6ZxCD8&ved=0ahUKEwjbkpes-Z3lAhVVL6YKHQsIC_UQ4dUDCAs&uact=5'.format(encText,encText))
        encText1 = learn[0]
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
             encText1=('{}%20{}'.format(encText1,learn[i]))
        html3 =('https://namu.wiki/w/{}'.format(encText1))
        naver = ('네이버_링크')
        tree = ('나무위키_링크')
        google = ('구글_링크')
        youtu = ('유튜브_링크')
        embed = discord.Embed(
                colour=discord.Colour.green()
 )
        embed.set_thumbnail(
            url='https://cdn.discordapp.com/attachments/621324300120490006/633613406150328320/155151515151.jpg')
        embed.add_field(name='네이버,구글,나무위키,유튜브사이트로 검색할게~!', value='{}를 검색해볼게~\n-------------------------------------------------------'.format(life), inline=False)
        embed.add_field(name='네이버', value='[%s](<%s>)' % (naver,html1), inline=False)
        embed.add_field(name='구글', value='[%s](<%s>)' % (google, html2), inline=False)
        embed.add_field(name='나무위키', value='[%s](<%s>)' % (tree, html3), inline=False)
        embed.add_field(name='유튜브', value='[%s](<%s>)' %(youtu, html4), inline=False)
        embed.set_footer(
            icon_url='https://cdn.discordapp.com/attachments/621324300120490006/631109608638906408/12121212121211111.jpg',
            text=f'{message.author.name}아 우리 쉬는시간을 가지는건 어때??')
        await client.send_message(message.channel, embed=embed)
    if message.content.startswith('운세'):
        Point = random.randint(1,100)
        if Point >= 95:
            await client.send_message(message.channel,"{}의 운세는 '대길'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel, embed=discord.Embed(colour=discord.Colour.gold(), title='오늘 도박이라도 해보는게 어때?ㅋㅋ'))
        elif Point >= 85:
            await client.send_message(message.channel,"{}의 운세는 '길'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel, embed=discord.Embed(colour=discord.Colour.gold(), title='흐음 오늘 좋은 일이 생길 수도??'))
        elif Point >= 75:
            await client.send_message(message.channel,"{}의 운세는 '중길'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel, embed=discord.Embed(colour=discord.Colour.gold(), title='흐음 오늘 좋은 일이 생길 수도??'))
        elif Point >= 65:
            await client.send_message(message.channel,"{}의 운세는 '소길'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel, embed=discord.Embed(colour=discord.Colour.gold(), title=' 나름..? 나쁘지 않네'))
        elif Point >= 45:
            await client.send_message(message.channel,"{}의 운세는 '평'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel,
                                      embed=discord.Embed(colour=discord.Colour.gold(), title='평..범..해...시시하다'))
        elif Point >= 35:
            await client.send_message(message.channel,"{}의 운세는 '흉'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel,
                                      embed=discord.Embed(colour=discord.Colour.gold(), title=' 헤에에..'))
        elif Point >= 25:
            await client.send_message(message.channel,"{}의 운세는 '소흉'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel,
                                      embed=discord.Embed(colour=discord.Colour.gold(), title='오늘 발밑 보고 조심히 다녀~ㅋㅋ'))
        elif Point >= 15:
            await client.send_message(message.channel,"{}의 운세는 '반흉'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel,
                                      embed=discord.Embed(colour=discord.Colour.gold(),
                                                          title='당신은 반흉이요!!! 위이위이 저리가시오!'))
        elif Point >= 5:
            await client.send_message(message.channel,"{}의 운세는 '말흉'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel,
                                      embed=discord.Embed(colour=discord.Colour.gold(),
                                                          title='나한테 붙지마!! 흉 옮겠다...'))
        elif Point <= 5:
            await client.send_message(message.channel,"{}의 운세는 '대흉'이며,\n행운수치는 100/{}입니다.".format(message.author.name,Point))
            await client.send_message(message.channel,
                                  embed=discord.Embed(colour=discord.Colour.gold(),
                                                      title='너...오늘..집에 가만히..있어.. 좀 불안해~'))

    if message.content.startswith('이미지검색'):
        search = message.content[6:]
        html = requests.get('https://search.naver.com/search.naver?where=image&sm=tab_jum&query={}'.format(search))
        soup = bs(html.text, 'html.parser')
        imgs = soup.find('img', class_='_img')
        imgs2 = imgs.get('data-source')

        link = soup.find('div', {'class': 'photo_grid _box'})
        link2 = link.findAll('a')
        link3 = link2[1]
        link4 = link3.get('href')

        embed = discord.Embed(
            colour=discord.Colour.green()
        )
        embed.add_field(name='검색: '+search, value='이미지링크 : ' + link4, inline=False)
        embed.set_footer(icon_url='https://cdn.discordapp.com/attachments/621324300120490006/631109608638906408/12121212121211111.jpg',text=f'{message.author.name}선배가 요청하신 자료에요~ 저 잘했죠?!')
        embed.set_image(url=imgs2)  # 이미지의 링크를 지정해 이미지를 설정합니다.
        await client.send_message(message.channel, embed=embed)  # 메시지를 보냅니다.
    if message.content.startswith("확률"):
        split = message.content[3:]
        await client.send_message(message.channel,"1부터 10까지 좋아하는 숫자를 하나 골라줘")
        a = random.randint(1, 100)
        b = random.randint(1, 100)
        c = random.randint(1, 100)
        d = random.randint(1, 100)
        e = random.randint(1, 100)
        f = random.randint(1, 100)
        g = random.randint(1, 100)
        h = random.randint(1, 100)
        i = random.randint(1, 100)
        j = random.randint(1, 100)
        persent = [a,b,c,d,e,f,g,h,i,j]
        msg = await client.wait_for_message(author=message.author)
        if msg.content == '1':
            await client.send_message(message.channel,"너가 1를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(split, persent[0]))
        elif msg.content == '2':
            await client.send_message(message.channel,
                                      "너가 2를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[1]))
        elif msg.content == '3':
            await client.send_message(message.channel,
                                      "너가 3를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[2]))
        elif msg.content == '4':
            await client.send_message(message.channel,
                                      "너가 4를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[3]))
        elif msg.content == '5':
            await client.send_message(message.channel,
                                      "너가 5를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[4]))
        elif msg.content == '6':
            await client.send_message(message.channel,
                                      "너가 2를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[5]))
        elif msg.content == '7':
            await client.send_message(message.channel,
                                      "너가 7를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[6]))
        elif msg.content == '8':
            await client.send_message(message.channel,
                                      "너가 8를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[7]))
        elif msg.content == '9':
            await client.send_message(message.channel,
                                      "너가 9를 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[8]))
        elif msg.content == '10':
            await client.send_message(message.channel,
                                      "너가 10을 고른 결과로 {}의 확률 결과 {}%가 나왔어\n너가 다른 것을 골랐을때의 표를 출력할까?\n출력할려면 print라 입력해줘".format(
                                          split, persent[9]))
        await client.wait_for_message(author=message.author, content='print')
        await client.send_message(message.channel,
                                  '1번째: %d퍼, 2번째: %d퍼, 3번째: %d퍼, 4번째: %d퍼,\n5번재: %d퍼, 6번째: %d퍼, 7번째: %d퍼, 8번째: %d퍼\n9번째: %d퍼, 10번째: %d퍼' % (
                                  a, b, c, d, e, f, g, h, i, j))
    if message.content.startswith("메이플지지"):
        abritudy = message.content.split(" ")
        namereal = abritudy[1]
        htmll = requests.get('https://maple.gg/u/{}'.format(namereal))
        maplegg = 'https://maple.gg/'
        soup = bs(htmll.text, 'html.parser')
        maple = soup.find('section', {'class': 'container'})
        maple1 = maple.find('div', {'class': 'col-lg-8'})
        maple2 = maple1.find('img')['alt']  #서버
        maple21 = maple1.find('b').text #이름
        maple3 = maple1.find('ul')
        maple4 = maple3.find('li').text #레벨
        maple5 = maple3.findAll('li')
        maple6 = maple5[1].text #직업
        maple7 = maple5[2]
        maple8 = maple7.findAll('span')
        maple9 = maple8[1].text #인기도
        maple10 = maple1.find('div', {'class': 'row'})
        try:
            maple11 = maple10.find('a').text  # 길드
        except AttributeError:
            maple11 = 'X'
            print('길드없음')
        maple12 = maple10.find('span').text
        maple13 = maple12.split("\n")
        maple14 = maple13[0] #종합랭킹
        maple15 = maple10.findAll('span')
        maple16 = maple15[1].text #월드랭킹
        maple17 = maple15[2].text #직업랭킹
        picture = maple.find('img')
        picture1 = picture.get('src')
        embed = discord.Embed(colour=discord.Colour.green())
        embed.set_thumbnail(url=picture1)
        embed.add_field(name='메이플지지로 검색중~~',
                        value='{}를 검색해볼게~\n-------------------------------------------------------'.format(namereal),
                        inline=False)
        embed.add_field(name='검색결과', value='시버: %s\n닉네임: %s\n%s\n직업: %s\n인기도: %s\n길드: %s\n종합랭킹:%s\n월드랭킹:%s\n직업랭킹(월드): %s' % (maple2, maple21,maple4,maple6,maple9,maple11,maple14,maple16,maple17),inline=False)
        embed.add_field(name='자세히는 직접 검색해줘~', value='%s' % (maplegg), inline=False)
        embed.set_footer(
            icon_url='https://cdn.discordapp.com/attachments/621324300120490006/631109608638906408/12121212121211111.jpg',
            text=f'{message.author.name}너?! 메이플 경력좀 있는데?!')
        await client.send_message(message.channel, embed=embed)
    if message.content.startswith("재획"):
        await client.send_message(message.channel,"경쿠30분 4개: 1번\n기타는 2번을 입력시켜줘~")
        msg = await client.wait_for_message(author=message.author)

        if msg.content == '1' or msg.content == '1번':
            a0 =await client.send_message(message.channel,"{0.author.mention}재물획득의 비약을 시작중~!!".format(message))
            ax = await client.send_message(message.channel,"경험치 쿠폰 30분 시작되었어~ㅎ 4/1")
            await asyncio.sleep(5)
            await client.delete_message(ax)
            await asyncio.sleep(1790)
            a = await client.send_message(message.channel, "{0.author.mention}아 경험치 구폰을 곧 먹어야해 4/2".format(message))
            await asyncio.sleep(5)
            await client.delete_message(a)
            await asyncio.sleep(1795)
            b = await client.send_message(message.channel, "{0.author.mention}아 경험치 구폰을 곧 먹어야해 4/3".format(message))
            await asyncio.sleep(5)
            await client.delete_message(b)
            await asyncio.sleep(1795)
            c = await client.send_message(message.channel, "{0.author.mention}아 경험치 구폰을 곧 먹어야해 4/4".format(message))
            await asyncio.sleep(5)
            await client.delete_message(c)
             await asyncio.sleep(1800)
            await client.send_message(message.channel, "수고했어~")
            await client.delete_message(a0)
            if msg.content == '2' or msg.content == '2번':
                 await client.send_message(message.channel, "개발중")
    if message.content.startswith("도와줘") or message.content.startswith("help"):
            channel = message.channel
            embed = discord.Embed(
                title="내가 할 수 있는 일들의 목록이야!",
                description="""잘 읽고 나한테 말해줘!! ------------------------------------------------------------""",
                colour=discord.Colour.blue()
 )

            embed.set_thumbnail(
                url='https://media.discordapp.net/attachments/405331598922219521/608627874299248661/151551155151511515.jpg')

            embed.add_field(name='!융', value='!융 [질문] 종은 것 싫은것을 답해줍니다' , inline=False)
            embed.add_field(name='주사위', value='1~6의 숫자를 골라줍니다', inline=False)
            embed.add_field(name='골라', value='골라 [고를꺼1] [고를꺼2] [고를꺼3] ...', inline=False)
            embed.add_field(name='청소', value='메세지를 삭제해줍니다. | 청소 (개수)', inline=False)
            embed.add_field(name='타이머', value='타이머 기능을 작동시킵니다 | 타이머 (초)', inline=False)
            embed.add_field(name='날씨', value='현재와 내일 오전,오후날씨를 알 수 있습니다. | 날씨 (지역)', inline=False)
            embed.add_field(name='저장', value='저장 {a} {b} a단어를 들으면 b를 말해줍니다', inline=False)
            embed.add_field(name='말해', value='말해 (융에게 시킬 단어)를 쓰면 융이 말해줌', inline=False)
            embed.add_field(name='운세', value='당신의 오늘의 운을 말해줍니다.', inline=False)
            embed.add_field(name='기억데이터', value='융이 저장한 단어들을 표시함', inline=False)
            embed.add_field(name='이미지검색', value='네이버이미지검색 첫번째 이미지를 가져옵니다 | 이미지검색 (검색어)', inline=False)
            embed.add_field(name='조용', value='메이플안하는사람을 조용히 시킵니다 | 조용 (맨션)', inline=False)
            embed.add_field(name='출근', value='융이 출근하여 다시 말을 할 수 있습니다', inline=False)
            embed.add_field(name='퇴근', value='융이 퇴근하여 사내가 조용합니다.', inline=False)
            await client.send_message(channel, embed=embed)

access_token=os.environ["bot_token"]
client.run(access_token)
